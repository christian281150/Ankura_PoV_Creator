from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook

NORMALISE_DIR = Path(__file__).parents[1] / "py" / "normalise"
sys.path.insert(0, str(NORMALISE_DIR))

from segments import extract_segments, parse_segment_sheet
from sheet_classifier import classify_workbook


FIXTURE = Path(__file__).parent / "fixtures" / "Textilkontor_Walter_Seidensticker_GmbH_Co_KG_Bielefeld_Konzernabschluss_FY2025_model.xlsx"


def test_fixture_extracts_product_and_geographic_values_in_eur() -> None:
    workbook = load_workbook(FIXTURE, read_only=True, data_only=True)
    extracted = extract_segments(workbook, classify_workbook(workbook))

    hemden = next(item for item in extracted.figures if item.segment_type == "product" and item.segment_name == "Hemden" and item.fiscal_year == 2016)
    inland = next(item for item in extracted.figures if item.segment_type == "geography" and item.segment_name == "Umsatzerlöse Inland" and item.fiscal_year == 2025)

    assert hemden.value == 145_608_000
    assert hemden.unit == "EUR"
    assert hemden.presentation_basis == "bruttoumsatzerloese"
    assert inland.value == 65_116_000
    assert inland.provenance.sheet.startswith("FY2025_")
    assert any("conflicting duplicate disclosure" in review.reason for review in extracted.reviews)


def test_percentage_table_is_not_coerced_and_unknown_basis_is_reviewed() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FY2025_percentages"
    worksheet.append(["", "2024/2025"])
    worksheet.append(["", "%"])
    worksheet.append(["A. Segmente"])
    worksheet.append(["1. Hemden", "55,5%"])

    extracted = parse_segment_sheet(worksheet)

    assert len(extracted.figures) == 1
    figure = extracted.figures[0]
    assert figure.value == 55.5
    assert figure.unit == "PCT"
    assert figure.metric == "revenue_share"
    assert figure.presentation_basis is None
    assert figure.flags == ("presentation_basis_unknown",)
    assert extracted.reviews


def test_dropped_segment_is_flagged_without_backfilling() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "FY2025_table"
    worksheet.append(["", "2024/2025"])
    worksheet.append(["", "T€", "T€"])
    worksheet.append(["A. Segmente"])
    worksheet.append(["1. Hemden", 100])
    worksheet.append(["2. Mieterträge", 1])
    worksheet.append(["3. Bruttoumsatzerlöse", 101])
    prior = workbook.create_sheet("FY2023_table")
    prior.append(["", "2022/2023"])
    prior.append(["", "T€"])
    prior.append(["A. Segmente"])
    prior.append(["1. Hemden", 90])
    prior.append(["2. Mieterträge", 1])
    prior.append(["3. Bruttoumsatzerlöse", 91])
    classifications = {worksheet.title: "anhang_umsatzsplit", prior.title: "anhang_umsatzsplit"}

    extracted = extract_segments(workbook, classifications)

    assert [item.fiscal_year for item in extracted.figures if item.segment_name == "Mieterträge"] == [2023, 2025]
    assert "product:Mieterträge:revenue not disclosed for FY2024; no continuity assumed" in extracted.discontinuities
