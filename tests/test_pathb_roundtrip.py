"""Regression tests for Lane D (Path B): template <-> producer round trip.

Proving ground: the real FY2024 Seidensticker filing, consolidated via
Path A (extractor.consolidate) and reconciled via py/series into a real
EntitySeries -- not a hand-built one. That series is written into the Path B
input format and read back, proving Path B is the same contract from a
different entry point, not a parallel path.

No line items need excluding from the round-trip source: Konzernbilanzverlust
(the one Path-A-emitted std_id that used to have no formal taxonomy entry --
consolidate.py produced it via an inline _SUBTOTAL_EXTENSIONS exact-text-match
shortcut, id "PL_GKV-BILANZVERLUST", which lib.hgb_map.by_id() never
recognised) is now a real taxonomy entry, "BS-P.A.KG-IV" (Lane A follow-up).
Path A no longer emits the old shortcut id at all.
"""
from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

import pytest

from contract.models import EntitySeries, FiscalYearEnd
from extractor.consolidate import build_multi_year_tables
from series.reconcile import build_entity_series

from acquire.pathb.producer import PathBValidationError, produce_entity_series
from acquire.pathb.schema import (
    COL_CLIENT_LABEL, COL_CURRENCY, COL_FRAMEWORK, COL_METHOD_FLAG,
    COL_PNL_METHOD, COL_PRESENTATION_BASIS, COL_SCOPE_FLAG, COL_STD_ID, COL_UNIT,
    COMPANY_FIELD_COL, COMPANY_SHEET, COMPANY_VALUE_COL, FIELD_ENTITY_ID,
    FIELD_FY_END_DAY, FIELD_FY_END_MONTH, LINE_ITEMS_SHEET, METADATA_COLUMNS,
)
from acquire.pathb.template import write_blank_template, write_filled_template

FY2024_FIXTURE = Path(__file__).parent / "fixtures" / "seidensticker_extracted_tables.json"
ENTITY_ID = "HRA-8217-AG-Bielefeld"

_COMPARED_FIELDS = ("framework", "pnl_method", "unit", "currency", "presentation_basis", "scope_flag", "method_flag")


def _path_a_series() -> EntitySeries:
    tables = build_multi_year_tables(json.loads(FY2024_FIXTURE.read_text(encoding="utf-8-sig")))
    return build_entity_series(ENTITY_ID, FiscalYearEnd(month=4, day=30), [tables])


def test_path_a_proving_fixture_has_a_realistic_number_of_line_items() -> None:
    """Sanity check on the proving ground itself, not the round trip."""
    series = _path_a_series()
    assert len(series.line_items) > 30


def test_round_trip_preserves_std_ids_values_and_declared_metadata() -> None:
    original = _path_a_series()
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "filled.xlsx"
        write_filled_template(original, xlsx_path)
        round_tripped = produce_entity_series(xlsx_path)

    assert {li.std_id for li in round_tripped.line_items} == {li.std_id for li in original.line_items}
    assert round_tripped.fiscal_years == original.fiscal_years
    assert round_tripped.fiscal_year_end == original.fiscal_year_end

    original_by_id = {li.std_id: li for li in original.line_items}
    round_tripped_by_id = {li.std_id: li for li in round_tripped.line_items}
    for std_id, line_item in original_by_id.items():
        rt_line_item = round_tripped_by_id[std_id]
        rt_points_by_fy = {point.fy: point for point in rt_line_item.points}
        for point in line_item.points:
            rt_point = rt_points_by_fy.get(point.fy)
            assert rt_point is not None, f"{std_id} FY{point.fy} missing after round trip"
            assert rt_point.observations[0].value == point.observations[0].value, (
                f"{std_id} FY{point.fy}: {point.observations[0].value} -> {rt_point.observations[0].value}"
            )
            for field in _COMPARED_FIELDS:
                assert getattr(rt_point, field) == getattr(point, field), (
                    f"{std_id} FY{point.fy}.{field}: {getattr(point, field)!r} -> {getattr(rt_point, field)!r}"
                )


def test_round_trip_provenance_and_source_kind_legitimately_differ() -> None:
    """Not a bug: Path B's own observations are genuinely user-supplied, not
    filing-sourced, and the series itself is genuinely a user_workbook, not a
    filings series -- these SHOULD differ from Path A's original values."""
    original = _path_a_series()
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "filled.xlsx"
        write_filled_template(original, xlsx_path)
        round_tripped = produce_entity_series(xlsx_path)

    assert original.source_kind == "filings"
    assert round_tripped.source_kind == "user_workbook"

    sample = round_tripped.line_items[0].points[0].observations[0]
    assert sample.provenance.kind == "user_supplied"
    original_sample = original.line_items[0].points[0].observations[0]
    assert original_sample.provenance.kind == "filing"


def test_producer_refuses_a_workbook_missing_a_required_declaration() -> None:
    """The all-or-nothing gate: one row missing presentation_basis must
    refuse the ENTIRE workbook, not silently drop that one row and emit
    everything else."""
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "incomplete.xlsx"
        _write_minimal_workbook(
            xlsx_path,
            company={FIELD_ENTITY_ID: "test-co", FIELD_FY_END_MONTH: 12, FIELD_FY_END_DAY: 31},
            rows=[
                {
                    COL_CLIENT_LABEL: "Revenue", COL_STD_ID: "PL_GKV-1", COL_FRAMEWORK: "ifrs",
                    COL_PNL_METHOD: "ukv", COL_UNIT: "EUR", COL_CURRENCY: "EUR",
                    COL_PRESENTATION_BASIS: "",  # missing -- the deliberate defect
                    "2024": 1_000_000,
                },
            ],
        )
        with pytest.raises(PathBValidationError) as excinfo:
            produce_entity_series(xlsx_path)

    assert any("presentation_basis" in problem for problem in excinfo.value.problems)
    # No partial series, no attribute to inspect -- the exception itself is the proof nothing was emitted.


def test_producer_refuses_an_unrecognised_std_id() -> None:
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "bad_std_id.xlsx"
        _write_minimal_workbook(
            xlsx_path,
            company={FIELD_ENTITY_ID: "test-co", FIELD_FY_END_MONTH: 12, FIELD_FY_END_DAY: 31},
            rows=[
                {
                    COL_CLIENT_LABEL: "Not a real concept", COL_STD_ID: "PL_GKV-DOES-NOT-EXIST",
                    COL_FRAMEWORK: "hgb", COL_PNL_METHOD: "gkv", COL_UNIT: "EUR", COL_CURRENCY: "EUR",
                    COL_PRESENTATION_BASIS: "n/a", "2024": -54_181_477.90,
                },
            ],
        )
        with pytest.raises(PathBValidationError) as excinfo:
            produce_entity_series(xlsx_path)

    assert any("not a recognised taxonomy id" in problem for problem in excinfo.value.problems)


def test_producer_refuses_the_same_std_id_declared_twice_for_one_year() -> None:
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "duplicate.xlsx"
        common = {
            COL_STD_ID: "PL_GKV-1", COL_FRAMEWORK: "ifrs", COL_PNL_METHOD: "ukv",
            COL_UNIT: "EUR", COL_CURRENCY: "EUR", COL_PRESENTATION_BASIS: "umsatzerloese",
        }
        _write_minimal_workbook(
            xlsx_path,
            company={FIELD_ENTITY_ID: "test-co", FIELD_FY_END_MONTH: 12, FIELD_FY_END_DAY: 31},
            rows=[
                {**common, COL_CLIENT_LABEL: "Revenue (system A)", "2024": 1_000_000},
                {**common, COL_CLIENT_LABEL: "Revenue (system B)", "2024": 1_000_001},
            ],
        )
        with pytest.raises(PathBValidationError) as excinfo:
            produce_entity_series(xlsx_path)

    assert any("more than one row" in problem for problem in excinfo.value.problems)


def test_blank_template_has_the_expected_sheets_and_columns() -> None:
    with TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "blank.xlsx"
        write_blank_template(xlsx_path)
        import openpyxl
        workbook = openpyxl.load_workbook(xlsx_path)

    assert workbook.sheetnames == [COMPANY_SHEET, LINE_ITEMS_SHEET]
    company_header = [cell.value for cell in next(workbook[COMPANY_SHEET].iter_rows(min_row=1, max_row=1))]
    assert company_header == [COMPANY_FIELD_COL, COMPANY_VALUE_COL]
    items_header = [cell.value for cell in next(workbook[LINE_ITEMS_SHEET].iter_rows(min_row=1, max_row=1))]
    assert items_header == list(METADATA_COLUMNS)


def _write_minimal_workbook(path: Path, company: dict[str, Any], rows: list[dict[str, Any]]) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    company_sheet = workbook.active
    company_sheet.title = COMPANY_SHEET
    company_sheet.append([COMPANY_FIELD_COL, COMPANY_VALUE_COL])
    for field, value in company.items():
        company_sheet.append([field, value])

    items_sheet = workbook.create_sheet(LINE_ITEMS_SHEET)
    years = sorted({key for row in rows for key in row if key not in METADATA_COLUMNS})
    items_sheet.append(list(METADATA_COLUMNS) + years)
    for row in rows:
        items_sheet.append([row.get(column, "") for column in METADATA_COLUMNS] + [row.get(year, "") for year in years])

    workbook.save(path)
