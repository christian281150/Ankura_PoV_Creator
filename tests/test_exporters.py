"""Regression test for Lane F: the ALL-GuV export sheet, the human review
surface an analyst actually looks at to sanity-check the mapper's work.

Before reproducing anything, this lane's own instruction was to verify the
original diagnosis ("Umsatzerloese blank for every year") still holds
against the current pipeline rather than assume it. It doesn't, fully: Lane
A already fixed the underlying non-resolution that caused a true blank.
What reproduces now is different -- Umsatzerloese IS present, but as an
unformatted text string, not a number, because _as_number's predecessor
(_parse_num_cell) assumed every cell was raw PDF-extracted German-locale
text. A multi-year overview table's cells (consolidate.py's
build_multi_year_tables output) are already-parsed Python floats; re-parsing
str(103152036.57) as if "." were a German thousands separator silently
returns None, so the value fell through to the plain-text branch. That
still fails the spirit of "non-empty for every year" for a human review
surface even though the cell technically wasn't blank -- an unformatted,
left-aligned text string reads as broken to an analyst scanning the sheet.
"""
from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Any

from extractor.consolidate import build_multi_year_tables
from extractor.exporters import _as_number, export_to_excel_v2

FY2024_FIXTURE = Path(__file__).parent / "fixtures" / "seidensticker_extracted_tables.json"


def _extracted_tables() -> list[dict[str, Any]]:
    return json.loads(FY2024_FIXTURE.read_text(encoding="utf-8-sig"))


def _all_guv_sheet(tmp_path: Path):
    import openpyxl

    result = build_multi_year_tables(_extracted_tables())
    out_path = tmp_path / "export.xlsx"
    export_to_excel_v2(result, {"company": "Test", "doc_type": "FY", "fy": "2024"}, out_path=out_path)
    workbook = openpyxl.load_workbook(out_path)
    sheet_name = next(name for name in workbook.sheetnames if "GuV" in name)
    return workbook[sheet_name]


def _row_by_label_prefix(sheet, prefix: str):
    for row in sheet.iter_rows(min_row=6):
        label = str(row[0].value or "")
        if label.startswith(prefix):
            return row
    raise AssertionError(f"no row starting with {prefix!r} found")


def test_umsatzerloese_is_a_non_empty_number_for_every_year() -> None:
    with TemporaryDirectory() as tmp_dir:
        sheet = _all_guv_sheet(Path(tmp_dir))
        row = _row_by_label_prefix(sheet, "1. Umsatzerl")
        year_cells = row[1:]
        assert year_cells, "expected at least one fiscal-year column"
        for cell in year_cells:
            assert cell.value not in (None, ""), f"{cell.coordinate} is blank"
            assert isinstance(cell.value, (int, float)), (
                f"{cell.coordinate} is {cell.value!r} ({type(cell.value).__name__}), "
                "not a number -- an analyst reviewing this sheet would see unformatted text"
            )
            assert cell.number_format != "General", f"{cell.coordinate} has no numeric format applied"


def test_blank_label_subtotal_rows_are_also_numeric_and_bold() -> None:
    """Gesamtleistung (blank raw_label, see Lane A) is exactly the kind of
    row _as_number's predecessor bug also silently broke: its value fell
    through to text, and the bold/border subtotal formatting that depends on
    detecting a numeric value in a blank-label row was lost with it."""
    with TemporaryDirectory() as tmp_dir:
        sheet = _all_guv_sheet(Path(tmp_dir))
        subtotal_row = next(
            row for row in sheet.iter_rows(min_row=6)
            if row[0].value in (None, "") and row[1].value not in (None, "")
        )
        for cell in subtotal_row[1:]:
            assert isinstance(cell.value, (int, float)), f"{cell.coordinate}: {cell.value!r}"
            assert cell.font.bold is True
            assert cell.border.top.style is not None


def test_as_number_still_parses_raw_german_locale_text() -> None:
    """The original use case (per-year raw tables extracted straight from
    the PDF, before consolidation) must still work: those cells are strings
    in German locale, not Python floats."""
    assert _as_number("103.152.036,57") == 103152036.57
    assert _as_number("-8.833.400,55") == -8833400.55
    assert _as_number("4.8") is None  # an Anhang cross-reference, not an amount
    assert _as_number("") is None
    assert _as_number(None) is None


def test_as_number_passes_through_an_already_parsed_float() -> None:
    assert _as_number(103152036.57) == 103152036.57
    assert _as_number(-8833400.55) == -8833400.55
