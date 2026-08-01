"""Workbook/CSV exporters and the mapping-audit sheet derivation."""

import re
from pathlib import Path
from typing import Optional
from rich.prompt import Prompt

from ._core import _HGB_AVAILABLE, _hgb, _parse_num_cell, console, sanitize_filename
from .extract import _classify_table


def export_to_csv(tables: list[dict], result: dict, out_dir: Optional[Path] = None) -> int:
    """
    Export each table to a CSV file using pandas.
    Returns number of tables exported.
    """
    company  = sanitize_filename(result["company"])
    doc_type = sanitize_filename(result["doc_type"])
    fy       = result["fy"]
    base_dir = out_dir or Path.cwd()
    exported = 0

    for t in tables:
        filename = f"{company}_{doc_type}_{fy}_table{t['index']}.csv"
        out_path = base_dir / filename

        for attempt in range(1, 3):
            try:
                import pandas as pd
                df = pd.DataFrame(t["rows"])
                df.to_csv(str(out_path), index=False, header=False, encoding="utf-8-sig")
                console.print(f"[green]Exported: {out_path.resolve()}[/green]")
                exported += 1
                break
            except OSError as exc:
                console.print(f"[red]Cannot write to {base_dir}: {exc}[/red]")
                alt = Prompt.ask("Enter an alternative output path")
                base_dir = Path(alt.strip())
                base_dir.mkdir(parents=True, exist_ok=True)
                out_path = base_dir / filename

    return exported


def _acct_indent(text: str) -> int:
    """
    Return the Excel indentation level (0-4) for a description-column cell.

    Hierarchy used in German Bilanz / GuV / KFR:
      0  top-level totals: Aktiva, Passiva, Summe …, Bilanzsumme
      1  capital-letter sections: A. B. C.
      2  Roman-numeral sub-sections: I. II. III. IV. V. …
      3  Arabic numbers 1. 2. or lowercase  a) b)
      4  double-letter  aa) bb)  or  davon / darunter notes

    IMPORTANT: Roman numerals (I, V, X prefix) must be checked BEFORE the
    generic capital-letter pattern, otherwise "I." matches as level 1.
    """
    t = (text or "").strip()
    if not t:
        return 0

    # Level 0 — aggregate totals and main headers
    _L0 = re.compile(
        r"^(Aktiva|Passiva|Bilanzsumme|Eigenkapital$"
        r"|Summe\b|Gesamt\b|Jahresuber|Jahresüber"
        r"|Ergebnis\b|Gewinn\b|Verlust\b)",
        re.IGNORECASE,
    )
    if _L0.match(t):
        return 0

    # Level 4 — deepest: double-letter aa) or davon/darunter notes
    if re.match(r"^[a-z]{2,}\)", t) or re.match(r"^(davon|darunter)\b", t, re.IGNORECASE):
        return 4

    # Level 3 — Arabic numerals  1.  or lowercase single-letter  a)
    if re.match(r"^\d+\.\s", t) or re.match(r"^[a-z]\)", t):
        return 3

    # Level 2 — Roman numerals  I.  II.  III.  IV.  V.  IX.  …
    # Must come BEFORE the capital-letter check so "I." is not caught as level 1
    if re.match(r"^[IVX]+\.\s", t):
        return 2

    # Level 1 — capital-letter sections  A.  B.  C.
    if re.match(r"^[A-Z]\.\s", t):
        return 1

    return 0


def _acct_bold(text: str) -> bool:
    """
    Return True when the description-column cell should be bold in Excel.

    Bold: top-level aggregate rows and main capital-letter sections (A. B. C.).
    NOT bold: Roman-numeral sub-sections (I. II. III.), Arabic / lowercase items.

    IMPORTANT: Roman numeral check must precede the capital-letter check so
    "I. Immaterielle …" is NOT marked bold.
    """
    t = (text or "").strip()
    if not t:
        return False

    # Never bold: Roman numerals, Arabic, lowercase, double-letter
    if re.match(r"^[IVX]+\.\s", t):
        return False
    if re.match(r"^\d+\.\s", t):
        return False
    if re.match(r"^[a-z]", t):
        return False

    # Bold: Aktiva / Passiva / Bilanzsumme / Summe … / Gesamt …
    if re.match(
        r"^(Aktiva|Passiva|Bilanzsumme|Eigenkapital$"
        r"|Summe\b|Gesamt\b|Jahresuber|Jahresüber"
        r"|Ergebnis\b|Gewinn\b|Verlust\b)",
        t, re.IGNORECASE,
    ):
        return True

    # Bold: capital-letter sections  A.  B.  C.
    if re.match(r"^[A-Z]\.\s", t):
        return True

    return False


def export_to_excel(tables: list[dict], result: dict,
                    out_path: Optional[Path] = None,
                    decimal_sep: str = ".",
                    thousand_sep: str = ",",
                    pdf_dir: Optional[Path] = None) -> tuple:
    """
    Export all tables to a single Excel workbook, one sheet per table.
    Each sheet begins with 4 metadata rows (title, export date, source, input
    path) followed by a blank row, then the financial table itself.
    Returns (table_count, saved_path).
    """
    import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _THIN_TOP   = Border(top=Side(border_style="thin"))
    _MEDIUM_TOP = Border(top=Side(border_style="medium"))

    # Metadata values computed once per workbook
    export_date = _dt.datetime.now().strftime("%d/%m/%Y")
    input_path  = str(pdf_dir.resolve()) if pdf_dir else "—"
    META_ROWS   = 5   # title + date + source + input + blank spacer
    # Row offsets (1-based Excel rows)
    R_TITLE  = 1
    R_DATE   = 2
    R_SOURCE = 3
    R_INPUT  = 4
    R_BLANK  = 5
    # Table data starts at row 6

    company  = sanitize_filename(result["company"])
    doc_type = sanitize_filename(result["doc_type"])
    fy       = result["fy"]

    if out_path is None:
        out_path = Path.cwd() / f"{company}_{doc_type}_{fy}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL   = PatternFill("solid", fgColor="1F538D")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT     = Font(name="Calibri", size=11)
    TITLE_FONT    = Font(name="Calibri", size=13, bold=True, color="1F538D")
    META_FONT     = Font(name="Calibri", size=10, color="475569")
    CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=False)

    used_names: set[str] = set()

    for t in tables:
        heading  = " ".join((t.get("heading") or "").split())
        label    = t.get("doc_label", "")
        if t.get("multi_year"):
            raw_name = re.sub(r'[\\/*?:\[\]]', "", heading)[:31] or "Overview"
        else:
            prefix   = f"{label}_" if label else ""
            raw_name = re.sub(r'[\\/*?:\[\]]', "", f"{prefix}{heading}")[:28] or f"Table {t['index']}"
        name = raw_name
        n = 2
        while name in used_names:
            suffix = f" ({n})"
            name = raw_name[:31 - len(suffix)] + suffix
            n += 1
        used_names.add(name)

        ws = wb.create_sheet(title=name)
        rows = t.get("rows", [])
        if not rows:
            continue

        max_cols = max(len(r) for r in rows)

        # ── Metadata header (rows 1-5) ────────────────────────────────────
        if t.get("multi_year"):
            _tn   = {0: "Bilanz", 1: "GuV", 2: "Kapitalflussrechnung"}.get(_classify_table(t), "Overview")
            _ystr = "  |  ".join(str(y) for y in t.get("years", []))
            sheet_title = f"{_tn}  —  Multi-Year Overview  ({_ystr})" if _ystr else heading
        else:
            sheet_title = f"{label}  —  {heading}" if label else heading

        # Merge each metadata row across all columns for a clean look
        for mr in range(1, META_ROWS + 1):
            ws.merge_cells(start_row=mr, start_column=1,
                           end_row=mr, end_column=max(max_cols, 1))

        c = ws.cell(row=R_TITLE,  column=1, value=sheet_title)
        c.font = TITLE_FONT

        c = ws.cell(row=R_DATE,   column=1, value=f"Export date: {export_date}")
        c.font = META_FONT

        c = ws.cell(row=R_SOURCE, column=1, value="Source: Unternehmensregister")
        c.font = META_FONT

        c = ws.cell(row=R_INPUT,  column=1, value=f"Input data: {input_path}")
        c.font = META_FONT

        ws.row_dimensions[R_TITLE].height  = 20
        ws.row_dimensions[R_BLANK].height  = 6    # narrow spacer

        # ── Financial table (rows 6+) ─────────────────────────────────────
        # Pre-scan: identify subtotal rows (empty description + numeric values)
        subtotal_set: set[int] = set()
        for ri0, row0 in enumerate(rows):
            if ri0 == 0:
                continue
            padded0 = (list(row0) + [""] * max_cols)[:max_cols]
            desc0   = str(padded0[0] if padded0 else "").strip()
            if not desc0 and any(
                _parse_num_cell(v, thousand_sep=".", decimal_sep=",") is not None
                for v in padded0[1:]
            ):
                subtotal_set.add(ri0)
        last_subtotal = max(subtotal_set) if subtotal_set else -1

        for ri, row in enumerate(rows, 1):
            xr          = ri + META_ROWS          # actual Excel row (offset by metadata)
            padded      = (list(row) + [""] * max_cols)[:max_cols]
            is_subtotal = (ri - 1) in subtotal_set
            is_grandtot = is_subtotal and (ri - 1) == last_subtotal

            for ci, val in enumerate(padded, 1):
                raw = val if val is not None else ""
                if ri > 1 and ci > 1:
                    num = _parse_num_cell(raw, thousand_sep=".", decimal_sep=",")
                    if num is not None:
                        cell = ws.cell(row=xr, column=ci, value=num)
                        cell.font = Font(name="Calibri", size=11, bold=is_subtotal)
                        _int = (abs(num - round(num)) < 1e-9)
                        cell.number_format = '#,##0' if _int else '#,##0.00'
                        if is_grandtot:
                            cell.border = _MEDIUM_TOP
                        elif is_subtotal:
                            cell.border = _THIN_TOP
                        continue
                cell = ws.cell(row=xr, column=ci,
                               value=str(raw) if raw != "" else "")
                if ri == 1:
                    cell.font      = HEADER_FONT
                    cell.fill      = HEADER_FILL
                    cell.alignment = CENTER_ALIGN
                elif ci == 1:
                    cell.font      = Font(name="Calibri", size=11,
                                          bold=_acct_bold(str(raw)) or is_subtotal)
                    cell.alignment = Alignment(indent=_acct_indent(str(raw)))
                    if is_grandtot:
                        cell.border = _MEDIUM_TOP
                    elif is_subtotal:
                        cell.border = _THIN_TOP
                else:
                    cell.font = BODY_FONT
                    if is_grandtot:
                        cell.border = _MEDIUM_TOP
                    elif is_subtotal:
                        cell.border = _THIN_TOP

        # Auto-size columns (cap at 50)
        for ci in range(1, max_cols + 1):
            col_letter = get_column_letter(ci)
            max_width  = max(
                (len(str(ws.cell(row=xr, column=ci).value or ""))
                 for xr in range(1, ws.max_row + 1)),
                default=8,
            )
            ws.column_dimensions[col_letter].width = min(max_width + 3, 50)

        # Freeze below the column-header row (row 6 = META_ROWS + 1)
        ws.freeze_panes = f"A{META_ROWS + 2}"

    wb.save(str(out_path))
    console.print(f"[green]Excel saved: {out_path.resolve()}[/green]")
    return len(tables), out_path


def export_to_excel_v2(tables: list,
                       result: dict,
                       out_path: "Optional[Path]" = None,
                       decimal_sep: str = ".",
                       thousand_sep: str = ",",
                       pdf_dir: "Optional[Path]" = None,
                       all_tables: "Optional[list]" = None,
                       review_meta: "Optional[list]" = None) -> tuple:
    """
    Redesign-era exporter.  Produces a complete workbook:

      • One sheet per OVERVIEW (multi-year) table        — ALL-Bilanz, ALL-GuV, ALL-KFR
      • One sheet per individual per-year (raw) table    — as extracted
      • One "Mapping Audit" sheet
            columns: raw_label, std_id, canonical_en, match_type, fiscal_year, company

    The original export_to_excel() is left untouched; this is an additive function.

    Args
    ----
    tables       : OVERVIEW (multi_year) tables to write first.
    result       : a representative filing dict (company/doc_type/fy) for naming.
    all_tables   : every per-year table (raw extractions) for the raw sheets +
                   audit rows.  Falls back to *tables* when None.
    review_meta  : optional pre-computed audit rows (list of dicts with keys
                   raw_label, std_id, canonical_en, match_type, fiscal_year,
                   company).  When None the audit sheet is derived via hgb_map.
    """
    import datetime as _dt
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    _THIN_TOP   = Border(top=Side(border_style="thin"))
    _MEDIUM_TOP = Border(top=Side(border_style="medium"))

    overview_tables = [t for t in tables if t.get("multi_year")]
    raw_tables      = [t for t in (all_tables if all_tables is not None else tables)
                       if not t.get("multi_year")]

    export_date = _dt.datetime.now().strftime("%d/%m/%Y")
    input_path  = str(pdf_dir.resolve()) if pdf_dir else "—"
    META_ROWS   = 5
    R_TITLE, R_DATE, R_SOURCE, R_INPUT, R_BLANK = 1, 2, 3, 4, 5

    company  = sanitize_filename(result.get("company", "export")) if result else "export"
    doc_type = sanitize_filename(result.get("doc_type", "")) if result else ""
    fy       = result.get("fy", "") if result else ""

    if out_path is None:
        out_path = Path.cwd() / f"{company}_{doc_type}_{fy}_overview.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL  = PatternFill("solid", fgColor="1F538D")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT    = Font(name="Calibri", size=11)
    TITLE_FONT   = Font(name="Calibri", size=13, bold=True, color="1F538D")
    META_FONT    = Font(name="Calibri", size=10, color="475569")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

    used_names: set = set()

    def _unique_name(raw: str) -> str:
        """Excel-safe, ≤31-char, de-duplicated worksheet name."""
        raw = re.sub(r'[\\/*?:\[\]]', "", raw)[:31] or "Sheet"
        name, n = raw, 2
        while name in used_names:
            suffix = f" ({n})"
            name = raw[:31 - len(suffix)] + suffix
            n += 1
        used_names.add(name)
        return name

    def _write_table_sheet(t: dict, sheet_name: str, sheet_title: str):
        """Write one table as a worksheet (meta header rows + column headers + data)."""
        ws   = wb.create_sheet(title=sheet_name)
        rows = t.get("rows", [])
        if not rows:
            return
        max_cols = max(len(r) for r in rows)

        for mr in range(1, META_ROWS + 1):
            ws.merge_cells(start_row=mr, start_column=1,
                           end_row=mr, end_column=max(max_cols, 1))
        ws.cell(row=R_TITLE,  column=1, value=sheet_title).font = TITLE_FONT
        ws.cell(row=R_DATE,   column=1, value=f"Export date: {export_date}").font = META_FONT
        ws.cell(row=R_SOURCE, column=1, value="Source: Unternehmensregister").font = META_FONT
        ws.cell(row=R_INPUT,  column=1, value=f"Input data: {input_path}").font = META_FONT
        ws.row_dimensions[R_TITLE].height = 20
        ws.row_dimensions[R_BLANK].height = 6

        subtotal_set: set = set()
        for ri0, row0 in enumerate(rows):
            if ri0 == 0:
                continue
            padded0 = (list(row0) + [""] * max_cols)[:max_cols]
            desc0   = str(padded0[0] if padded0 else "").strip()
            if not desc0 and any(
                _parse_num_cell(v, thousand_sep=".", decimal_sep=",") is not None
                for v in padded0[1:]):
                subtotal_set.add(ri0)
        last_subtotal = max(subtotal_set) if subtotal_set else -1

        for ri, row in enumerate(rows, 1):
            xr          = ri + META_ROWS
            padded      = (list(row) + [""] * max_cols)[:max_cols]
            is_subtotal = (ri - 1) in subtotal_set
            is_grandtot = is_subtotal and (ri - 1) == last_subtotal
            for ci, val in enumerate(padded, 1):
                raw = val if val is not None else ""
                if ri > 1 and ci > 1:
                    num = _parse_num_cell(raw, thousand_sep=".", decimal_sep=",")
                    if num is not None:
                        cell = ws.cell(row=xr, column=ci, value=num)
                        cell.font = Font(name="Calibri", size=11, bold=is_subtotal)
                        _int = (abs(num - round(num)) < 1e-9)
                        cell.number_format = '#,##0' if _int else '#,##0.00'
                        if is_grandtot:   cell.border = _MEDIUM_TOP
                        elif is_subtotal: cell.border = _THIN_TOP
                        continue
                cell = ws.cell(row=xr, column=ci, value=str(raw) if raw != "" else "")
                if ri == 1:
                    cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER_ALIGN
                elif ci == 1:
                    cell.font = Font(name="Calibri", size=11,
                                     bold=_acct_bold(str(raw)) or is_subtotal)
                    cell.alignment = Alignment(indent=_acct_indent(str(raw)))
                    if is_grandtot:   cell.border = _MEDIUM_TOP
                    elif is_subtotal: cell.border = _THIN_TOP
                else:
                    cell.font = BODY_FONT
                    if is_grandtot:   cell.border = _MEDIUM_TOP
                    elif is_subtotal: cell.border = _THIN_TOP

        for ci in range(1, max_cols + 1):
            col_letter = get_column_letter(ci)
            max_width  = max((len(str(ws.cell(row=xr, column=ci).value or ""))
                              for xr in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[col_letter].width = min(max_width + 3, 50)
        ws.freeze_panes = f"A{META_ROWS + 2}"

    # ── 1) OVERVIEW sheets ────────────────────────────────────────────────
    for t in overview_tables:
        heading = " ".join((t.get("heading") or "").split())
        tn      = {0: "Bilanz", 1: "GuV", 2: "Kapitalflussrechnung"}.get(
                    _classify_table(t), "Overview")
        ystr    = "  |  ".join(str(y) for y in t.get("years", []))
        title   = f"{tn}  —  Multi-Year Overview  ({ystr})" if ystr else heading
        name    = _unique_name(heading or f"ALL-{tn}")
        _write_table_sheet(t, name, title)

    # ── 2) Per-year raw sheets ────────────────────────────────────────────
    for t in raw_tables:
        heading = " ".join((t.get("heading") or "").split())
        label   = t.get("doc_label", "")
        title   = f"{label}  —  {heading}" if label else heading
        prefix  = f"{label}_" if label else ""
        name    = _unique_name(f"{prefix}{heading}" or f"Table {t.get('index','')}")
        _write_table_sheet(t, name, title)

    # ── 3) Mapping Audit sheet ────────────────────────────────────────────
    audit = wb.create_sheet(title="Mapping Audit")
    audit_cols = ["raw_label", "std_id", "canonical_en",
                  "match_type", "fiscal_year", "company"]
    for ci, h in enumerate(audit_cols, 1):
        c = audit.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER_ALIGN

    audit_rows = review_meta if review_meta is not None else \
        _derive_audit_rows(raw_tables)
    for ri, rec in enumerate(audit_rows, 2):
        for ci, key in enumerate(audit_cols, 1):
            audit.cell(row=ri, column=ci, value=str(rec.get(key, "") or "")).font = BODY_FONT
    for ci, _ in enumerate(audit_cols, 1):
        letter = get_column_letter(ci)
        width  = max((len(str(audit.cell(row=r, column=ci).value or ""))
                      for r in range(1, audit.max_row + 1)), default=10)
        audit.column_dimensions[letter].width = min(width + 3, 60)
    audit.freeze_panes = "A2"

    wb.save(str(out_path))
    console.print(f"[green]Excel (v2) saved: {out_path.resolve()}[/green]")
    n_sheets = len(overview_tables) + len(raw_tables) + 1
    return n_sheets, out_path


def _derive_audit_rows(raw_tables: list) -> list:
    """Build Mapping-Audit rows from per-year tables using hgb_map.

    One row per (raw_label, fiscal_year, company).  std_id/canonical_en are
    filled only for unambiguous single-candidate matches; ambiguous / no-match
    labels are still listed (so they are auditable) with the match_type recorded.
    """
    out: list = []
    seen: set = set()
    for t in raw_tables:
        company = str(t.get("_company", "") or "")
        fy      = str(t.get("doc_label", "") or "")
        for row in (t.get("rows") or [])[1:]:
            if not row:
                continue
            desc = str(row[0] or "").strip()
            if not desc:
                continue
            sig = (desc, fy, company)
            if sig in seen:
                continue
            seen.add(sig)
            std_id = canon = ""
            mt = "none"
            if _HGB_AVAILABLE:
                try:
                    res   = _hgb.lookup(desc)
                    mt    = res.get("match_type", "none")
                    cands = res.get("candidates", [])
                    if len(cands) == 1:
                        std_id = cands[0].get("std_id", "")
                        canon  = cands[0].get("canonical_en", "")
                except Exception:
                    pass
            out.append({
                "raw_label": desc, "std_id": std_id, "canonical_en": canon,
                "match_type": mt, "fiscal_year": fy, "company": company,
            })
    return out
