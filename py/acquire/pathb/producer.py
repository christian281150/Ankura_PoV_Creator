"""Path B producer: read a filled-in PathB_Input_Template.xlsx and emit the
frozen contract.models.EntitySeries -- or refuse entirely.

All-or-nothing by construction: every problem across every row is collected
before anything is raised, and if there is even one, no EntitySeries is
returned at all -- never a partial one with the bad row silently dropped.

Not built here, by design (see docs/final-push-lanes.md, Lane D): saved
mappings per company and a refresh path for recurring (non-one-off)
engagements. That decision is still open; this module only ever reads one
filled workbook and produces one EntitySeries from it.

Also not enforced here, because the field does not exist on this module's
output: earnings_basis="adjusted" requiring user-supplied reconciliation
adjustments (AGENTS.md, Path B section) is a contract.models.ContentBlock
concern, not an EntitySeries/LineItemPoint one -- ContentBlock is what
py/blocks/ (not yet built) will eventually produce from this series. Whoever
builds py/blocks/ still needs to enforce that rule, for both Path A and
Path B content blocks; it is untouched here because the models this module
emits have no such field to violate.
"""
from __future__ import annotations

from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import lib.hgb_map as hgb_map
from contract.models import (
    EntitySeries,
    FiscalYearEnd,
    LineItemObservation,
    LineItemPoint,
    LineItemSeries,
    UserSuppliedSeriesProvenance,
)

from . import schema as s


class PathBValidationError(ValueError):
    """Raised when the input workbook is incomplete or invalid in any way.

    Carries every problem found, not just the first one encountered.
    """

    def __init__(self, problems: list[str]) -> None:
        self.problems = problems
        super().__init__("Path B input rejected:\n" + "\n".join(f"  - {p}" for p in problems))


def _year_columns(header: list[str]) -> dict[int, int]:
    """{fiscal_year: 0-based column index} for every plain-4-digit-year header cell."""
    years: dict[int, int] = {}
    for index, name in enumerate(header):
        text = name.strip()
        if text.isdigit() and len(text) == 4:
            years[int(text)] = index
    return years


def _cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def produce_entity_series(path: str | Path) -> EntitySeries:
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)

    missing_sheets = [name for name in (s.COMPANY_SHEET, s.LINE_ITEMS_SHEET) if name not in workbook.sheetnames]
    if missing_sheets:
        raise PathBValidationError([f"missing required sheet {name!r}" for name in missing_sheets])

    company: dict[str, Any] = {}
    for row in workbook[s.COMPANY_SHEET].iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        company[str(row[0]).strip()] = row[1]

    problems: list[str] = [
        f"Company sheet: missing required field {field!r}"
        for field in s.COMPANY_FIELDS if company.get(field) in (None, "")
    ]

    items_rows = list(workbook[s.LINE_ITEMS_SHEET].iter_rows(values_only=True))
    if not items_rows:
        raise PathBValidationError(problems + ["Line Items sheet is empty"])

    header = [str(cell or "").strip() for cell in items_rows[0]]
    missing_columns = [name for name in s.METADATA_COLUMNS if name not in header]
    if missing_columns:
        raise PathBValidationError(problems + [f"Line Items sheet: missing required column {name!r}" for name in missing_columns])

    col_index = {name: header.index(name) for name in s.METADATA_COLUMNS}
    year_cols = _year_columns(header)
    if not year_cols:
        problems.append("Line Items sheet: no fiscal-year value columns found (expected a plain 4-digit year header, e.g. \"2024\")")

    parsed_rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(items_rows[1:], start=2):
        metadata = {name: _cell(row, idx) for name, idx in col_index.items()}
        year_cells = {year: _cell(row, idx) for year, idx in year_cols.items()}
        if all(v in (None, "") for v in metadata.values()) and all(v in (None, "") for v in year_cells.values()):
            continue  # a genuinely blank template row -- not a partially-started one

        label = metadata.get(s.COL_CLIENT_LABEL) or f"row {row_number}"
        row_problems: list[str] = []

        for field in s.REQUIRED_LINE_ITEM_COLUMNS:
            if metadata.get(field) in (None, ""):
                row_problems.append(f"missing {field}")

        std_id = metadata.get(s.COL_STD_ID)
        if std_id and hgb_map.by_id(str(std_id)) is None:
            row_problems.append(f"std_id {std_id!r} is not a recognised taxonomy id")

        for field, valid in (
            (s.COL_FRAMEWORK, s.VALID_FRAMEWORKS), (s.COL_PNL_METHOD, s.VALID_PNL_METHODS),
            (s.COL_UNIT, s.VALID_UNITS), (s.COL_CURRENCY, s.VALID_CURRENCIES),
            (s.COL_PRESENTATION_BASIS, s.VALID_PRESENTATION_BASES),
        ):
            value = metadata.get(field)
            if value and value not in valid:
                row_problems.append(f"{field} {value!r} must be one of {valid}")

        year_values: dict[int, Decimal] = {}
        for year, cell in year_cells.items():
            if cell in (None, ""):
                continue
            try:
                year_values[year] = Decimal(str(cell))
            except InvalidOperation:
                row_problems.append(f"FY{year} value {cell!r} is not numeric")

        if row_problems:
            problems.extend(f"Line Items row {row_number} ({label!r}): {p}" for p in row_problems)
            continue

        parsed_rows.append({**metadata, "row_number": row_number, "year_values": year_values})

    if problems:
        raise PathBValidationError(problems)

    by_std_id: dict[str, list[dict[str, Any]]] = {}
    for row in parsed_rows:
        by_std_id.setdefault(str(row[s.COL_STD_ID]), []).append(row)

    merge_problems: list[str] = []
    line_items: list[LineItemSeries] = []
    for std_id, group_rows in sorted(by_std_id.items()):
        year_to_row: dict[int, dict[str, Any]] = {}
        for row in group_rows:
            for year in row["year_values"]:
                if year in year_to_row:
                    merge_problems.append(
                        f"std_id {std_id!r} has a value for FY{year} on more than one row "
                        f"(rows {year_to_row[year]['row_number']} and {row['row_number']}) -- "
                        "Path B is single-source; this is a data-entry conflict, not a restatement"
                    )
                else:
                    year_to_row[year] = row
        if merge_problems:
            continue
        points = [
            LineItemPoint(
                fy=year,
                unit=row[s.COL_UNIT],
                currency=row[s.COL_CURRENCY],
                framework=row[s.COL_FRAMEWORK],
                pnl_method=row[s.COL_PNL_METHOD],
                presentation_basis=row[s.COL_PRESENTATION_BASIS],
                scope_flag=row[s.COL_SCOPE_FLAG] or None,
                method_flag=row[s.COL_METHOD_FLAG] or None,
                observations=[LineItemObservation(
                    value=row["year_values"][year],
                    provenance=UserSuppliedSeriesProvenance(kind="user_supplied"),
                    restated=False,
                )],
                resolution=None,
            )
            for year, row in sorted(year_to_row.items())
        ]
        line_items.append(LineItemSeries(std_id=std_id, points=points))

    if merge_problems:
        raise PathBValidationError(merge_problems)

    return EntitySeries(
        entity_id=str(company[s.FIELD_ENTITY_ID]),
        source_kind="user_workbook",
        fiscal_year_end=FiscalYearEnd(month=int(company[s.FIELD_FY_END_MONTH]), day=int(company[s.FIELD_FY_END_DAY])),
        fiscal_years=sorted({year for row in parsed_rows for year in row["year_values"]}),
        line_items=line_items,
    )
