"""Write the Path B input template.

write_blank_template() produces the downloadable, empty workbook an analyst
fills in by hand -- the mapping screen's only concrete artefact, since no
GUI exists yet (see the Lane D design note in docs/final-push-lanes.md).

write_filled_template() writes an already-built EntitySeries into the same
format. Used to prove the round trip (Path A/py.series output -> template ->
producer -> same series back), not part of the analyst-facing flow.
"""
from __future__ import annotations

from pathlib import Path

from contract.models import EntitySeries, LineItemSeries

from . import schema as s

_UNIFORM_FIELDS = (
    "framework", "pnl_method", "unit", "currency",
    "presentation_basis", "scope_flag", "method_flag",
)


def write_blank_template(path: str | Path) -> None:
    """Write the blank, downloadable template with dropdown-assisted columns.

    Dropdowns are convenience only -- openpyxl data validation does not stop
    a pasted value that bypasses the dropdown, so producer.py re-validates
    every cell regardless and is the actual gate.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = openpyxl.Workbook()
    company = workbook.active
    company.title = s.COMPANY_SHEET
    company.append([s.COMPANY_FIELD_COL, s.COMPANY_VALUE_COL])
    for field in s.COMPANY_FIELDS:
        company.append([field, ""])

    items = workbook.create_sheet(s.LINE_ITEMS_SHEET)
    items.append(list(s.METADATA_COLUMNS))

    last_row = "1000"
    dropdowns = (
        (s.COL_FRAMEWORK, s.VALID_FRAMEWORKS),
        (s.COL_PNL_METHOD, s.VALID_PNL_METHODS),
        (s.COL_UNIT, s.VALID_UNITS),
        (s.COL_CURRENCY, s.VALID_CURRENCIES),
        (s.COL_PRESENTATION_BASIS, s.VALID_PRESENTATION_BASES),
    )
    for column_name, choices in dropdowns:
        column_letter = get_column_letter(s.METADATA_COLUMNS.index(column_name) + 1)
        validation = DataValidation(type="list", formula1='"' + ",".join(choices) + '"', allow_blank=True)
        items.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{last_row}")

    workbook.save(path)


def _uniform_metadata(line_item: LineItemSeries) -> dict[str, object]:
    """One declared value per field for this whole line item.

    The one-off Path B mapping model declares framework/pnl_method/unit/
    currency/presentation_basis/scope_flag/method_flag once per line item
    (per docs/final-push-lanes.md's Lane D note), not once per fiscal year,
    even though contract.models.LineItemPoint allows the latter. Raises if
    an input series genuinely varies one of these across years for the same
    std_id, rather than silently collapsing to whichever year came first.
    """
    values = {field: {getattr(point, field) for point in line_item.points} for field in _UNIFORM_FIELDS}
    inconsistent = [field for field, distinct in values.items() if len(distinct) > 1]
    if inconsistent:
        raise ValueError(
            f"{line_item.std_id}: {', '.join(inconsistent)} differs across fiscal years -- "
            "the one-off Path B template declares these once per line item; this series "
            "cannot be losslessly re-templated without a per-year override, which is not built"
        )
    return {field: next(iter(distinct)) for field, distinct in values.items()}


def write_filled_template(series: EntitySeries, path: str | Path) -> None:
    import openpyxl

    workbook = openpyxl.Workbook()
    company = workbook.active
    company.title = s.COMPANY_SHEET
    company.append([s.COMPANY_FIELD_COL, s.COMPANY_VALUE_COL])
    company.append([s.FIELD_ENTITY_ID, series.entity_id])
    company.append([s.FIELD_FY_END_MONTH, series.fiscal_year_end.month])
    company.append([s.FIELD_FY_END_DAY, series.fiscal_year_end.day])

    items = workbook.create_sheet(s.LINE_ITEMS_SHEET)
    years = series.fiscal_years
    items.append(list(s.METADATA_COLUMNS) + [str(year) for year in years])

    for line_item in series.line_items:
        metadata = _uniform_metadata(line_item)
        by_fy = {point.fy: point for point in line_item.points}
        # No client wording survives in an EntitySeries (canonical, std_id-keyed
        # by design) -- the std_id itself stands in as client_label here. Not
        # asserted on by the round-trip proof, which checks std_id/value/
        # declared-metadata equality, not this display convenience.
        row = [
            line_item.std_id, line_item.std_id,
            metadata["framework"], metadata["pnl_method"], metadata["unit"],
            metadata["currency"], metadata["presentation_basis"],
            metadata["scope_flag"] or "", metadata["method_flag"] or "",
        ]
        for year in years:
            point = by_fy.get(year)
            if point is None or not point.observations:
                row.append("")
                continue
            if len(point.observations) != 1:
                raise ValueError(
                    f"{line_item.std_id} FY{year} has {len(point.observations)} observations "
                    "(an unresolved restatement) -- write_filled_template only handles a "
                    "single-source series; Path A output with an open conflict is not valid "
                    "Path B input until that conflict is resolved upstream"
                )
            row.append(float(point.observations[0].value))
        items.append(row)

    workbook.save(path)
