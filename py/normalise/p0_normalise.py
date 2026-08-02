"""Deterministic P0 workbook normalisation.

Mappings are loaded from the Bundesanzeiger submodule.  A label maps only when
its normalised (display-number-free) form is an exact source/extension key;
otherwise it is written to ``py/normalise/reviews/unmapped_queue.csv``.
"""
from __future__ import annotations

import csv
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

NORMALISE_DIR = Path(__file__).resolve().parent
HGB_DATA_DIR = NORMALISE_DIR.parent / "acquire" / "bundesanzeiger" / "lib" / "hgb_data"
if str(HGB_DATA_DIR) not in sys.path:
    sys.path.insert(0, str(HGB_DATA_DIR))

from hgb_lookup_reference import HGBMapper, _normalize  # noqa: E402
from sheet_classifier import classify_workbook  # noqa: E402

_NUMBER = re.compile(r"^[+\-\s]*[\d.,]+$")
_LEADING_ITEM = re.compile(r"^\s*(?:[a-z]|[ivxlcdm]+|\d+)\s*[.)]\s*", re.I)

# These codes are controlled, visible extensions until the generated taxonomy
# includes its missing HGB heading rows.  They are not a copied source mapping.
REQUIRED_TAXONOMY_EXTENSIONS: dict[str, str] = {
    "materialaufwand": "PL_GKV-5",
    "personalaufwand": "PL_GKV-6",
    "abschreibungen": "PL_GKV-7",
    "gesamtleistung": "PL_GKV-GESAMTLEISTUNG",
    "rohergebnis": "PL_GKV-ROHERGEBNIS",
    "konzernbilanzverlust": "PL_GKV-BILANZVERLUST",
    "nichtdurchvermoegenseinlagengedeckterverlustanteil": "BS-P-NEGEQ",
    "veraenderungdesbestandsanunfertigenundfertigenerzeugnissen": "PL_GKV-2",
    "erhoehungoderverminderungdesbestandesanfertigenundunfertigenerzeugnissen": "PL_GKV-2",
}
CLIENT_ALIASES_PATH = NORMALISE_DIR / "aliases" / "client_aliases.csv"


def parse_de_number(value: Any) -> float | None:
    """Parse German strings, including ``'+ 1.914.645,32'``."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or not _NUMBER.fullmatch(text):
        return None
    negative = text.lstrip().startswith("-")
    text = text.lstrip("+- ").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif text.count(".") > 1:
        text = text.replace(".", "")
    try:
        parsed = float(text)
    except ValueError:
        return None
    return -parsed if negative else parsed


def _match_key(label: str) -> str:
    """Apply only presentation cleanup before the source normalisation."""
    cleaned = _LEADING_ITEM.sub("", label.strip())
    cleaned = re.sub(r"\s*\(GKV\)\s*$", "", cleaned, flags=re.I)
    return _normalize(cleaned)


class Mapper:
    """Exact-only mapper.  There is deliberately no substring/fuzzy fallback."""

    def __init__(self, mapping_dir: Path = HGB_DATA_DIR) -> None:
        self._source = HGBMapper(str(mapping_dir))
        self._aliases: dict[str, str] = {}
        if CLIENT_ALIASES_PATH.exists():
            with CLIENT_ALIASES_PATH.open(encoding="utf-8", newline="") as file:
                for row in csv.DictReader(file):
                    self._aliases[row["normalized_key"]] = row["std_id"]

    def lookup(self, label: str) -> tuple[str | None, str]:
        key = _match_key(label)
        if not key:
            return None, "none"
        code = REQUIRED_TAXONOMY_EXTENSIONS.get(key)
        if code:
            return code, "extension_exact"
        code = self._aliases.get(key)
        if code:
            return code, "client_alias_exact"
        code = self._source.labels.get(key)
        if code:
            return code, "exact"
        return None, "none"
    
    def is_subtotal(self, code: str | None) -> bool:
        """Extractor rule 2: subtotals must not enter the mapped actuals."""
        if not code:
            return False
        record = self._source.taxonomy.get(code)
        return bool(record and record.get("is_subtotal"))


def detect_unit(rows: list[list[Any]]) -> float:
    blob = " ".join(str(cell).lower() for row in rows[:10] for cell in row if cell is not None)
    return 1000.0 if any(token in blob for token in ("t€", "teuro", "teur", "tsd.")) else 1.0


def parse_year_header(value: Any) -> int | None:
    years = re.findall(r"20\d{2}", str(value or ""))
    return max(map(int, years)) if years else None


def find_year_blocks(rows: list[list[Any]], max_scan: int = 10) -> tuple[int | None, list[tuple[int, int, int]]]:
    """Return fiscal-year column blocks; values can occupy either column."""
    for row_index, row in enumerate(rows[:max_scan]):
        hits = [(column, parse_year_header(value)) for column, value in enumerate(row) if column and parse_year_header(value)]
        if not hits or len({year for _, year in hits}) != len(hits):
            continue
        blocks = []
        for index, (column, year) in enumerate(hits):
            end = hits[index + 1][0] - 1 if index + 1 < len(hits) else len(row) - 1
            blocks.append((year, column, end))
        return row_index, blocks
    return None, []


def _is_davon(label: str) -> bool:
    return _normalize(label).startswith("davon")


def extract_statement(ws: Any, mapper: Mapper) -> list[dict[str, Any]]:
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    header_index, blocks = find_year_blocks(rows)
    if header_index is None:
        return []
    multiplier = detect_unit(rows)
    records: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        label = str(row[0]).strip() if row and row[0] is not None else ""
        if not label or _is_davon(label):
            continue
        values: dict[int, float] = {}
        for year, start, end in blocks:
            for column in range(start, min(end + 1, len(row))):
                value = parse_de_number(row[column])
                if value is not None:
                    values[year] = value * multiplier
                    break
        if not values:
            continue
        std_id, match_type = mapper.lookup(label)
        is_sub = mapper.is_subtotal(std_id)
        records.append({
            "raw_label": label,
            "std_id": std_id,
            "match_type": match_type,
            "row_type": "subtotal" if is_sub else "line",
            "values": values,
            "unit": "EUR",
            "presentation_basis": "umsatzerloese" if std_id == "PL_GKV-1" else None,
            "provenance": {"doc": ws.title.split("_")[0], "sheet": ws.title, "row": row_number, "page": None},
        })
    return records


def merge_on_std_id(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    """Merge duplicate mapped lines by std_id and retain the first source per FY."""
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    unmapped: list[dict[str, Any]] = []
    for record in records:
        std_id = record["std_id"]
        if not std_id:
            unmapped.append(record)
            continue
        target = merged.setdefault(std_id, {**record, "values": {}, "provenance_by_fy": {}})
        for year, value in record["values"].items():
            if year not in target["values"]:
                target["values"][year] = value
                target["provenance_by_fy"][year] = record["provenance"]
    return list(merged.values()) + unmapped


def write_unmapped_queue(records: Iterable[dict[str, Any]], path: Path) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    unique: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for record in records:
        if record["std_id"] is None:
            unique.setdefault(_match_key(record["raw_label"]), record)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=("raw_label", "normalized_key", "sheet", "row"))
        writer.writeheader()
        for key, record in unique.items():
            writer.writerow({"raw_label": record["raw_label"], "normalized_key": key,
                             "sheet": record["provenance"]["sheet"], "row": record["provenance"]["row"]})
    return len(unique)


def build_series(records: Iterable[dict[str, Any]], std_id: str) -> dict[int, float]:
    for record in records:
        if record["std_id"] == std_id and record.get("row_type") != "subtotal":
            return dict(record["values"])
    return {}


def main(xlsx_path: str, mapping_dir: str, out_path: str) -> None:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    mapper = Mapper(Path(mapping_dir))
    classifications = classify_workbook(workbook)
    raw_records: list[dict[str, Any]] = []
    sheets_parsed: list[str] = []
    for sheet_name, classification in classifications.items():
        # The workbook's ALL/MAPPING audit tabs are derived views, not filings.
        if not sheet_name.startswith("FY"):
            continue
        if classification != "guv":
            continue
        records = extract_statement(workbook[sheet_name], mapper)
        if records:
            raw_records.extend(records)
            sheets_parsed.append(sheet_name)

    queue_length = write_unmapped_queue(raw_records, NORMALISE_DIR / "reviews" / "unmapped_queue.csv")
    records = merge_on_std_id(raw_records)
    revenue = build_series(records, "PL_GKV-1")
    inventory = build_series(records, "PL_GKV-2")
    other_income = build_series(records, "PL_GKV-4")
    reconciliation = []
    for year in sorted(set(revenue) | set(inventory) | set(other_income)):
        r, i, o = revenue.get(year), inventory.get(year), other_income.get(year)
        reconciliation.append({"fy": year, "umsatzerloese": r, "bestandsveraenderung": i,
                               "sonstige_betriebliche_ertraege": o,
                               "gesamtleistung": r + i + o if None not in (r, i, o) else None})
    mapped = sum(record["std_id"] is not None for record in raw_records)
    result = {"coverage": {"sheets_parsed": sheets_parsed, "rows_extracted": len(raw_records),
                             "rows_mapped": mapped, "map_rate": mapped / len(raw_records) if raw_records else 0,
                             "unmapped_queue_length": queue_length,
                             "sheet_classification": classifications},
              "reconciliation": reconciliation, "rows": records}
    Path(out_path).write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"map rate: {result['coverage']['map_rate']:.1%} ({mapped}/{len(raw_records)})")
    print(f"unmapped queue length: {queue_length}")


if __name__ == "__main__":
    main(*sys.argv[1:4])
